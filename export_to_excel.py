#!/usr/bin/env python3
"""
Script para exportar dados do CRM para Excel
Uso: python export_to_excel.py [--period month] [--token YOUR_JWT_TOKEN]
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import requests
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ Erro: Biblioteca necessária não encontrada: {e}")
    print("\n📦 Instale as dependências com:")
    print("pip install requests pandas openpyxl")
    sys.exit(1)


class CRMExporter:
    def __init__(self, api_url: str, token: str):
        self.api_url = api_url
        self.headers = {"Authorization": f"Bearer {token}"}

    def fetch_dashboard_data(self, period: str = "month"):
        """Busca dados do dashboard via API"""
        try:
            print(f"🔄 Buscando dados do período: {period}...")
            response = requests.get(
                f"{self.api_url}/dashboard/export",
                headers=self.headers,
                params={"period": period},
                timeout=30
            )
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"❌ Erro ao buscar dados: {e}")
            sys.exit(1)

    def create_excel_report(self, data: dict, output_file: str):
        """Cria relatório Excel com múltiplas abas"""
        print(f"📊 Criando relatório Excel: {output_file}")

        # Criar um writer do Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Aba 1: Resumo Geral
            self._create_summary_sheet(writer, data)

            # Aba 2: Leads
            if data.get('export', {}).get('leads'):
                self._create_leads_sheet(writer, data['export']['leads'])

            # Aba 3: Clientes
            if data.get('export', {}).get('clients'):
                self._create_clients_sheet(writer, data['export']['clients'])

            # Aba 4: Empresas
            if data.get('export', {}).get('companies'):
                self._create_companies_sheet(writer, data['export']['companies'])

            # Aba 5: Timeline
            if data.get('timeline'):
                self._create_timeline_sheet(writer, data['timeline'])

        # Aplicar formatação avançada
        self._apply_advanced_formatting(output_file)

        print(f"✅ Relatório criado com sucesso!")
        print(f"📁 Arquivo: {output_file}")

    def _create_summary_sheet(self, writer, data):
        """Cria aba de resumo com estatísticas principais"""
        stats = data.get('stats', {})
        totals = stats.get('totals', {})
        growth = stats.get('growth', {})
        conversion = stats.get('conversion', {})

        summary_data = {
            'Métrica': [
                'Total de Leads',
                'Total de Clientes',
                'Total de Empresas',
                'Total de Usuários',
                '',
                'Crescimento de Leads (%)',
                'Crescimento de Clientes (%)',
                'Crescimento de Empresas (%)',
                '',
                'Taxa de Conversão (%)',
                'Leads Qualificados',
                'Leads Perdidos',
                '',
                'Período do Relatório',
                'Data de Exportação',
            ],
            'Valor': [
                totals.get('leads', 0),
                totals.get('clientes', 0),
                totals.get('empresas', 0),
                totals.get('usuarios', 0),
                '',
                f"{growth.get('leads', 0):+.1f}%",
                f"{growth.get('clientes', 0):+.1f}%",
                f"{growth.get('empresas', 0):+.1f}%",
                '',
                f"{conversion.get('rate', 0):.1f}%",
                conversion.get('qualified', 0),
                conversion.get('lost', 0),
                '',
                data.get('export', {}).get('period', 'N/A').upper(),
                datetime.now().strftime('%d/%m/%Y %H:%M'),
            ]
        }

        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Resumo Geral', index=False)

    def _create_leads_sheet(self, writer, leads_data):
        """Cria aba com dados detalhados de leads"""
        if not leads_data:
            return

        df = pd.DataFrame(leads_data)

        # Renomear colunas para português
        column_mapping = {
            'id': 'ID',
            'nome': 'Nome',
            'email': 'E-mail',
            'telefone': 'Telefone',
            'status': 'Status',
            'origem': 'Origem',
            'valor_estimado': 'Valor Estimado',
            'empresa': 'Empresa',
            'segmento': 'Segmento',
            'responsavel': 'Responsável',
            'criado_em': 'Criado Em',
            'atualizado_em': 'Atualizado Em',
        }

        df = df.rename(columns=column_mapping)

        # Formatar datas
        for col in ['Criado Em', 'Atualizado Em']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.strftime('%d/%m/%Y %H:%M')

        # Formatar valores monetários
        if 'Valor Estimado' in df.columns:
            df['Valor Estimado'] = df['Valor Estimado'].apply(
                lambda x: f"R$ {float(x):,.2f}" if pd.notna(x) else 'N/A'
            )

        df.to_excel(writer, sheet_name='Leads', index=False)

    def _create_clients_sheet(self, writer, clients_data):
        """Cria aba com dados detalhados de clientes"""
        if not clients_data:
            return

        df = pd.DataFrame(clients_data)

        column_mapping = {
            'id': 'ID',
            'nome': 'Nome',
            'email': 'E-mail',
            'telefone': 'Telefone',
            'cargo': 'Cargo',
            'ativo': 'Ativo',
            'empresa': 'Empresa',
            'segmento': 'Segmento',
            'responsavel': 'Responsável',
            'criado_em': 'Criado Em',
            'atualizado_em': 'Atualizado Em',
        }

        df = df.rename(columns=column_mapping)

        for col in ['Criado Em', 'Atualizado Em']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.strftime('%d/%m/%Y %H:%M')

        if 'Ativo' in df.columns:
            df['Ativo'] = df['Ativo'].apply(lambda x: 'Sim' if x else 'Não')

        df.to_excel(writer, sheet_name='Clientes', index=False)

    def _create_companies_sheet(self, writer, companies_data):
        """Cria aba com dados detalhados de empresas"""
        if not companies_data:
            return

        df = pd.DataFrame(companies_data)

        column_mapping = {
            'id': 'ID',
            'nome': 'Nome',
            'cnpj': 'CNPJ',
            'segmento': 'Segmento',
            'ativo': 'Ativo',
            'telefone': 'Telefone',
            'email': 'E-mail',
            'site': 'Site',
            'total_leads': 'Total Leads',
            'total_clientes': 'Total Clientes',
            'criado_em': 'Criado Em',
            'atualizado_em': 'Atualizado Em',
        }

        df = df.rename(columns=column_mapping)

        for col in ['Criado Em', 'Atualizado Em']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col]).dt.strftime('%d/%m/%Y %H:%M')

        if 'Ativo' in df.columns:
            df['Ativo'] = df['Ativo'].apply(lambda x: 'Sim' if x else 'Não')

        df.to_excel(writer, sheet_name='Empresas', index=False)

    def _create_timeline_sheet(self, writer, timeline_data):
        """Cria aba com dados da timeline"""
        if not timeline_data:
            return

        df = pd.DataFrame({
            'Período': timeline_data.get('labels', []),
            'Leads': timeline_data.get('leads', []),
            'Clientes': timeline_data.get('clientes', []),
        })

        df.to_excel(writer, sheet_name='Timeline', index=False)

    def _apply_advanced_formatting(self, filename: str):
        """Aplica formatação avançada ao Excel"""
        wb = load_workbook(filename)

        # Cores
        header_fill = PatternFill(start_color="00FF88", end_color="00FF88", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Formatar cabeçalho
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Congelar primeira linha
            ws.freeze_panes = 'A2'

        wb.save(filename)


def main():
    parser = argparse.ArgumentParser(
        description='Exporta dados do CRM para Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  python export_to_excel.py --token seu_token_jwt
  python export_to_excel.py --period week --token seu_token_jwt
  python export_to_excel.py --period quarter --output relatorio_trimestral.xlsx --token seu_token_jwt

Períodos disponíveis: today, week, month, quarter, year, all
        """
    )

    parser.add_argument(
        '--api-url',
        default='http://localhost:3000/api/v1',
        help='URL da API (padrão: http://localhost:3000/api/v1)'
    )

    parser.add_argument(
        '--token',
        required=True,
        help='Token JWT para autenticação'
    )

    parser.add_argument(
        '--period',
        default='month',
        choices=['today', 'week', 'month', 'quarter', 'year', 'all'],
        help='Período para filtrar os dados (padrão: month)'
    )

    parser.add_argument(
        '--output',
        default=None,
        help='Nome do arquivo de saída (padrão: relatorio_crm_YYYYMMDD_HHMMSS.xlsx)'
    )

    args = parser.parse_args()

    # Definir nome do arquivo de saída
    if args.output:
        output_file = args.output
    else:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'relatorio_crm_{timestamp}.xlsx'

    # Executar export
    print("🚀 CRM Data Exporter")
    print("=" * 50)

    exporter = CRMExporter(args.api_url, args.token)
    data = exporter.fetch_dashboard_data(args.period)
    exporter.create_excel_report(data, output_file)

    print("\n" + "=" * 50)
    print(f"✨ Exportação concluída com sucesso!")
    print(f"📊 Abas criadas: Resumo Geral, Leads, Clientes, Empresas, Timeline")


if __name__ == '__main__':
    main()
